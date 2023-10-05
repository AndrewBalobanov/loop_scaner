import pathlib

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from pandas import DataFrame, ExcelWriter
import os


def dir_list(path):
    import os
    dir_list = []

    for dirpath, dirnames, filenames in os.walk(os.path.normpath(path)):
        for name in filenames:
            file = os.path.join(dirpath, name)
            if file.endswith('.xlsx'):
                dir_list.append(file)
    return dir_list


directory = os.path.dirname(os.path.abspath(__file__))
files = dir_list(directory)


for file in files:
    path = file
    filename = f'{pathlib.Path(file).stem}.xlsx'
    result = f'result_{filename}'

    new_list = []

    wb = load_workbook(path)
    sheets = wb.sheetnames

    for i in range(len(sheets)):
        ws = wb[sheets[i]]
        for column in range(30, 60):
            effect = ws.cell(row=3, column=column).value
            if effect:
                effect_description = ws.cell(row=2, column=column).value
                for row in range(5, 41):
                    reason = ws.cell(row=row, column=23).value
                    cross_point = ws.cell(row=row, column=column).value
                    if cross_point:
                        reason_description = ws.cell(row=row, column=24).value
                        new_series = {'Причина': reason,
                                      'Описание причины': reason_description,
                                      'Следствие': effect,
                                      'Описание следствия': effect_description
                                      }
                        new_list.append(new_series)

    df = DataFrame(new_list)
    to_write = {'Sheet': df}
    with ExcelWriter(path=result, engine='xlsxwriter') as wb:
        for sheet_name, frame in to_write.items():
            frame.to_excel(wb, sheet_name=sheet_name, index=False)
            sheet = wb.sheets[sheet_name]
            for column in frame:
                base_column_width = max(frame[column].astype(str).map(len).max(), len(column)) + 4
                column_width = 1.05 * base_column_width
                col_idx = frame.columns.get_loc(column)

                style = wb.book.add_format()
                style.set_align('left')
                style.set_align('vcenter')

                sheet.set_column(col_idx, col_idx, column_width, style)


    def top_row(path):
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment

        wb = load_workbook(path)
        sheets = wb.sheetnames

        for i in range(len(sheets)):
            ws = wb[sheets[i]]

            row_1 = ws[1]
            top_style = Font(bold=True, name='Calibre', size=10, underline='single')
            for cell in row_1:
                if cell.value:
                    cell.font = top_style
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.freeze_panes = 'A2'

            ws.auto_filter.ref = ws.dimensions

        wb.save(path)


    wb = load_workbook(result)
    ws = wb['Sheet']

    for row in ws.iter_rows():
        for cell in row:
            thins = Side(border_style='thin', color='000000')
            cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

    top_row(result)
